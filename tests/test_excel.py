import json
import random
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter


from tcomextetl.common.excel import SimpleExcelDataReader
from tcomextetl.common.utils import read_file

json_fpath = Path(__file__).parent / 'misc' / 'excel_data.json'
test_data = json.loads(read_file(json_fpath))


sheets_num = 5


@pytest.fixture(params=test_data)
def save_test_data_in_excel(request):
    """ Create excel file with three sheets. One for testing skiprows parameter,
        one for use_cols parameter, one is empty to test sheet_indexes parameter
    """

    data_dicts = request.param

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=True) as tf:
        tmp_fpath = tf.name

    # sheets to create
    sheets_range = range(0, sheets_num + 1)

    # calculate range of inserting cells
    cells_range = f'A1:{get_column_letter(len(data_dicts[0].values()))}{len(data_dicts)}'

    skip_rows = [0, random.randint(1, 5), random.randint(1, 5)]

    sheets = random.sample(sheets_range, len(skip_rows))
    sheets.sort()

    wb = Workbook()
    wb.remove(wb.active)
    for i in sheets_range:
        wb.create_sheet(title=str(i))

    cols = random.randint(1, 5)

    data_tuples = [tuple(d.values()) for d in data_dicts]

    for idx, sh_idx in enumerate(sheets):
        ws = wb.worksheets[sh_idx]
        rows = skip_rows[idx]
        for row in data_tuples:
            ws.append(row)

        ws.move_range(cells_range, rows=rows, cols=cols)

    wb.save(filename=tmp_fpath)

    use_cols = f'{get_column_letter(cols+1)}:{get_column_letter(cols + len(data_dicts[0].values()))}'
    params = {'sheet_indexes': sheets, 'skip_rows': skip_rows, 'use_cols': use_cols}

    return tmp_fpath, data_tuples * len(sheets), params


def test_simple_excel_read(save_test_data_in_excel):

    tmp_fpath, data_tuples, params = save_test_data_in_excel
    reader = SimpleExcelDataReader(tmp_fpath, **params)
    res = [data for data in reader]
    assert data_tuples == res
